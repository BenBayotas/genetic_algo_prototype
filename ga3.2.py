import random
from openpyxl import Workbook
from openpyxl.styles import Alignment


# Define the Department class to hold courses
class Department:
    def __init__(self, name, courses):
        self.name = name
        self.courses = courses


# Define the Course class to hold sections and subjects
class Course:
    def __init__(self, code, name, year_levels):
        self.code = code
        self.name = name
        self.year_levels = year_levels


# Define the Section class for each section in a course
class Section:
    def __init__(self, section_name):
        self.section_name = section_name


# Define the Subject class to hold details of each subject
class Subject:
    def __init__(self, code, name, duration, available_days, rooms, instructor, instructor_avail, num_students):
        self.code = code
        self.name = name
        self.duration = duration  # "1 hour and 30 mins", "3 hours", "5 hours"
        self.available_days = available_days  # e.g., ["Monday", "Thursday"]
        self.rooms = rooms  # List of available rooms
        self.instructor = instructor
        self.instructor_avail = instructor_avail  # Time slots available for the instructor
        self.num_students = num_students


# Define the Schedule class for managing and optimizing schedules
class Schedule:
    def __init__(self, subjects, sections):
        self.subjects = subjects
        self.sections = sections
        self.schedule = {}  # To hold the subject schedules

    # Initialize a random schedule for subjects
    def initialize(self):
        for section in self.sections:
            for subject in self.subjects:
                available_times = self.generate_time_slots(subject.duration)
                available_rooms = subject.rooms
                available_instructor_times = subject.instructor_avail

                # Randomly choose a time and room that fit both instructor and subject constraints
                time_slot = random.choice(list(set(available_times) & set(available_instructor_times)))
                room = random.choice(available_rooms)

                # Schedule the subject for each of its available days
                for day in subject.available_days:
                    self.schedule[(section.section_name, subject.code, day)] = (time_slot, room)

    # Generate possible time slots based on the duration of the subject
    def generate_time_slots(self, duration):
        start_time = 7 * 60  # Start at 7:00 AM in minutes
        end_time = 21 * 60  # End at 9:00 PM in minutes
        slots = []

        # Convert duration to minutes
        if duration == "1 hour and 30 mins":
            slot_duration = 90
        elif duration == "3 hours":
            slot_duration = 180
        elif duration == "5 hours":
            slot_duration = 300
        else:
            raise ValueError("Invalid duration")

        # Create time slots every 30 minutes
        current_time = start_time
        while current_time + slot_duration <= end_time:
            hour = current_time // 60
            minute = current_time % 60
            slots.append(f"{hour:02d}:{minute:02d}")
            current_time += 30

        return slots

    # Calculate the fitness based on the number of conflicts
    def calculate_fitness(self):
        conflicts = 0
        schedule_slots = {}

        # Check for conflicts in the schedule
        for (section, subject_code, day), (start_time, room) in self.schedule.items():
            subject = next(sub for sub in self.subjects if sub.code == subject_code)
            occupied_slots = self.get_occupied_slots(start_time, subject.duration)

            for slot in occupied_slots:
                key = (slot, day, room)
                if key in schedule_slots:
                    conflicts += 1
                schedule_slots[key] = (section, subject_code)

        return 1 / (1 + conflicts)

    # Get time slots occupied by a subject
    def get_occupied_slots(self, start_time, duration):
        time_slots = []
        hour, minute = map(int, start_time.split(':'))
        total_minutes = hour * 60 + minute

        if duration == "1 hour and 30 mins":
            duration_minutes = 90
        elif duration == "3 hours":
            duration_minutes = 180
        elif duration == "5 hours":
            duration_minutes = 300
        else:
            raise ValueError("Invalid duration")

        while duration_minutes > 0:
            hour = total_minutes // 60
            minute = total_minutes % 60
            time_slots.append(f"{hour:02d}:{minute:02d}")
            total_minutes += 30
            duration_minutes -= 30

        return time_slots

    # Randomly mutate a schedule by changing time or room
    def mutate(self):
        section, subject_code, day = random.choice(list(self.schedule.keys()))
        subject = next((s for s in self.subjects if s.code == subject_code), None)
        available_times = self.generate_time_slots(subject.duration)
        available_rooms = subject.rooms

        new_time_slot = random.choice(available_times)
        new_room = random.choice(available_rooms)

        self.schedule[(section, subject_code, day)] = (new_time_slot, new_room)


# Crossover operation to generate new offspring from parents
def crossover(parent1, parent2):
    crossover_point = random.randint(0, len(parent1.schedule) - 1)
    child1 = Schedule(parent1.subjects, parent1.sections)
    child2 = Schedule(parent2.subjects, parent2.sections)

    for i, key in enumerate(parent1.schedule.keys()):
        if i < crossover_point:
            child1.schedule[key] = parent1.schedule[key]
            child2.schedule[key] = parent2.schedule[key]
        else:
            child1.schedule[key] = parent2.schedule[key]
            child2.schedule[key] = parent1.schedule[key]

    return child1, child2


# Genetic Algorithm implementation to find the best schedule
def genetic_algorithm(subjects, sections, population_size=100, generations=1000):
    population = [Schedule(subjects, sections) for _ in range(population_size)]

    # Initialize the population
    for schedule in population:
        schedule.initialize()

    # Run the genetic algorithm over generations
    for generation in range(generations):
        population = sorted(population, key=lambda x: x.calculate_fitness(), reverse=True)
        next_generation = population[:population_size // 2]

        while len(next_generation) < population_size:
            parent1 = random.choice(population[:population_size // 2])
            parent2 = random.choice(population[:population_size // 2])
            child1, child2 = crossover(parent1, parent2)

            if random.random() < 0.1:
                child1.mutate()
                child2.mutate()

            next_generation.extend([child1, child2])

        population = next_generation

    return max(population, key=lambda x: x.calculate_fitness())


# Export the schedule to an Excel file with merged cells
def export_to_excel(schedule, filename="schedule.xlsx"):
    time_slots = [f"{hour:02d}:{minute:02d}" for hour in range(7, 21) for minute in (0, 30)]
    days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']

    wb = Workbook()
    ws = wb.active
    ws.title = "Class Schedule"

    # Create headers
    ws.append(["Time"] + days)

    for i, time_slot in enumerate(time_slots, start=2):
        ws[f"A{i}"] = time_slot

    # Fill in the schedule
    for (section, subject_code, day), (start_time, room) in schedule.schedule.items():
        subject = next((s for s in schedule.subjects if s.code == subject_code), None)
        start_row = time_slots.index(start_time) + 2
        occupied_slots = schedule.get_occupied_slots(start_time, subject.duration)
        end_row = start_row + len(occupied_slots) - 1
        col = days.index(day) + 2
        cell_value = f"{section} - {subject_code} ({room})"

        # Merge cells according to the duration of the subject
        ws.merge_cells(start_row=start_row, start_column=col, end_row=end_row, end_column=col)
        top_left_cell = ws.cell(row=start_row, column=col)
        top_left_cell.value = cell_value
        top_left_cell.alignment = Alignment(horizontal='center', vertical='center')

    wb.save(filename)
    print(f"Schedule saved to {filename}")


# Example usage
subjects = [
    Subject("CS101", "Intro to Programming", "1 hour and 30 mins", ["Monday", "Thursday"], ["Room 1", "Room 2"], "Prof. A", ["08:00", "10:00"], 30),
    Subject("CS102", "Data Structures", "3 hours", ["Tuesday", "Friday"], ["Room 1", "Room 3"], "Prof. B", ["09:00", "13:00"], 25),
    Subject("CS103", "Algorithms", "5 hours", ["Wednesday"], ["Room 2", "Room 4"], "Prof. C", ["11:00"], 20),
]

sections = [Section("CS11"), Section("CS12"), Section("CS13")]

best_schedule = genetic_algorithm(subjects, sections)
export_to_excel(best_schedule)
