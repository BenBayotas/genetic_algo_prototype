import random
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment


# Subject class representing each subject's details
class Subject:
    def __init__(self, code, name, duration, days, room_avail, instructor_avail, num_students):
        self.code = code
        self.name = name
        self.duration = duration  # Duration: "1 hour and 30 mins", "3 hours", "5 hours"
        self.days = days
        self.room_avail = room_avail
        self.instructor_avail = instructor_avail
        self.num_students = num_students


# Schedule class that manages the creation and optimization of the schedule
class Schedule:
    def __init__(self, subjects):
        self.subjects = subjects
        self.schedule = {}

    # Initializes the schedule with random assignments of time and room slots
    def initialize(self):
        for subject in self.subjects:
            available_times = self.generate_time_slots(subject.duration)
            available_rooms = subject.room_avail
            time_slot = random.choice(available_times)
            room = random.choice(available_rooms)

            # Assign the subject to the same room and time across all its available days
            for day in subject.days:
                self.schedule[(subject.code, day)] = (time_slot, room)

    # Generates all possible starting time slots for a subject based on its duration
    def generate_time_slots(self, duration):
        start_time = 7 * 60  # 7:00 AM in minutes
        end_time = 21 * 60  # 9:00 PM in minutes
        slots = []

        # Duration in minutes
        if duration == "1 hour and 30 mins":
            slot_duration = 90
        elif duration == "3 hours":
            slot_duration = 180
        elif duration == "5 hours":
            slot_duration = 300
        else:
            raise ValueError("Invalid duration")

        # Generate all starting time slots within operating hours
        current_time = start_time
        while current_time + slot_duration <= end_time:
            hour = current_time // 60
            minute = current_time % 60
            slots.append(f"{hour:02d}:{minute:02d}")
            current_time += 30  # Increment in 30-minute intervals

        return slots

    # Calculates the fitness score based on the number of conflicts in the schedule
    def calculate_fitness(self):
        conflicts = 0
        schedule_slots = {}

        # Count conflicts in the schedule
        for (subject_code, day), (start_time, room) in self.schedule.items():
            duration = next(sub.duration for sub in self.subjects if sub.code == subject_code)
            occupied_slots = self.get_occupied_slots(start_time, duration)

            for slot in occupied_slots:
                key = (slot, day, room)
                if key in schedule_slots:
                    conflicts += 1
                schedule_slots[key] = subject_code

        return 1 / (1 + conflicts)

    # Gets all the time slots occupied by a subject based on its start time and duration
    def get_occupied_slots(self, start_time, duration):
        time_slots = []
        hour, minute = map(int, start_time.split(':'))
        total_minutes = hour * 60 + minute

        if duration == "1 hour and 30 mins":
            duration_minutes = 90
        elif duration == "3 hours":
            duration_minutes = 180
        elif duration == "5 hours":
            duration_minutes = 300
        else:
            raise ValueError("Invalid duration")

        # Generate all time slots occupied by the subject
        while duration_minutes > 0:
            hour = total_minutes // 60
            minute = total_minutes % 60
            time_slots.append(f"{hour:02d}:{minute:02d}")
            total_minutes += 30
            duration_minutes -= 30

        return time_slots

    # Mutates the schedule by randomly changing a subject's room or time
    def mutate(self):
        subject_code, day = random.choice(list(self.schedule.keys()))
        subject = next((s for s in self.subjects if s.code == subject_code), None)
        available_times = self.generate_time_slots(subject.duration)
        available_rooms = subject.room_avail

        new_time_slot = random.choice(available_times)
        new_room = random.choice(available_rooms)

        # Apply the change to all days the subject is scheduled
        for d in subject.days:
            self.schedule[(subject_code, d)] = (new_time_slot, new_room)


# Performs crossover between two parent schedules to produce offspring
def crossover(parent1, parent2):
    crossover_point = random.randint(0, len(parent1.schedule) - 1)
    child1 = Schedule(parent1.subjects)
    child2 = Schedule(parent2.subjects)

    for i, key in enumerate(parent1.schedule.keys()):
        if i < crossover_point:
            child1.schedule[key] = parent1.schedule[key]
            child2.schedule[key] = parent2.schedule[key]
        else:
            child1.schedule[key] = parent2.schedule[key]
            child2.schedule[key] = parent1.schedule[key]

    return child1, child2


# Main genetic algorithm function to optimize the schedule
def genetic_algorithm(subjects, population_size=100, generations=1000):
    population = [Schedule(subjects) for _ in range(population_size)]

    # Initialize each schedule in the population
    for schedule in population:
        schedule.initialize()

    # Run the algorithm for the specified number of generations
    for generation in range(generations):
        population = sorted(population, key=lambda x: x.calculate_fitness(), reverse=True)
        next_generation = population[:population_size // 2]

        # Generate new offspring
        while len(next_generation) < population_size:
            parent1 = random.choice(population[:population_size // 2])
            parent2 = random.choice(population[:population_size // 2])
            child1, child2 = crossover(parent1, parent2)

            if random.random() < 0.1:
                child1.mutate()
                child2.mutate()

            next_generation.extend([child1, child2])

        population = next_generation

    # Return the best schedule
    return max(population, key=lambda x: x.calculate_fitness())


# Exports the optimized schedule to an Excel file with merged cells based on duration
def export_to_excel(schedule, filename="schedule.xlsx"):
    time_slots = [f"{hour:02d}:{minute:02d}" for hour in range(7, 21) for minute in (0, 30)]
    days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']

    wb = Workbook()
    ws = wb.active
    ws.title = "Class Schedule"

    # Set up the headers for the time slots and days
    ws.append(["Time"] + days)

    for i, time_slot in enumerate(time_slots, start=2):
        ws[f"A{i}"] = time_slot

    # Fill the schedule in the Excel sheet and merge cells based on duration
    for (subject_code, day), (start_time, room) in schedule.schedule.items():
        subject = next((s for s in schedule.subjects if s.code == subject_code), None)
        start_row = time_slots.index(start_time) + 2
        occupied_slots = schedule.get_occupied_slots(start_time, subject.duration)
        end_row = start_row + len(occupied_slots) - 1
        col = days.index(day) + 2
        cell_value = f"{subject_code} ({room})"

        # Merge cells based on the subject's duration
        ws.merge_cells(start_row=start_row, start_column=col, end_row=end_row, end_column=col)
        ws.cell(row=start_row, column=col, value=cell_value).alignment = Alignment(horizontal='center',
                                                                                   vertical='center')

    wb.save(filename)
    print(f"Schedule saved to {filename}")


# Example usage with sample subjects
subjects = [

    Subject("GE119", "Living in an IT Era (with MOS Certificate)", "1 hour and 30 mins", ["Tuesday", "Friday"], ["CB 222", "CB 223", "CB 224", "CB 225", "CB 226", "CB 227"], ["Instructor B"],40),
    Subject("CS471", "Principles of Operating System", "3 hours", ["Monday", "Thursday"], ["CB 222", "CB 223", "CB 224", "CB 225", "CB 226", "CB 227"], ["Instructor C"], 20),
    Subject("CS470", "Research (CS Thesis Writing 2)", "1 hour and 30 mins", ["Tuesday", "Friday"], ["CB 305", "CB 304"], ["Instructor D"], 35),
    Subject("CB472-EL4", "Computational Science (with Certification for AI Professionals-Part 2)", "5 hours", ["Wednesday"], ["CB 305", "CB 304"], ["Instructor E"], 20),
    Subject("CS172", "Probability and Statistics", "1 hour and 30 mins", ["Monday", "Thursday"], ["CB 321", "CB 322", "CB 323", "CB 324", "CB 325", "CB 326", "CB 327"], ["Instructor A"], 30),
    # Add more subjects as needed
]

# Run the genetic algorithm and export the best schedule
best_schedule = genetic_algorithm(subjects)
export_to_excel(best_schedule)

